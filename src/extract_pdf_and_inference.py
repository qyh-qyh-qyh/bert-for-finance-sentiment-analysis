import os
import pandas as pd
import pdfplumber
import openpyxl
import re

import torch
import datasets
from transformers import AutoTokenizer, AutoModelForSequenceClassification

import configparser
config = configparser.ConfigParser()
config.read('config.ini')

pretrained_model_name_or_path=config.get("inference","pretrained_model_name_or_path")
pretrained_tokenizer_name_or_path=config.get("inference","pretrained_tokenizer_name_or_path")
trained_model_weights_path=config.get("inference","trained_model_weights_path")
dir_path=config.get("inference","dir_path")
excel_path=config.get("inference","excel_path")
ws_title=config.get("inference","ws_title")

model=AutoModelForSequenceClassification.from_pretrained(pretrained_model_name_or_path)
tokenizer=AutoTokenizer.from_pretrained(pretrained_tokenizer_name_or_path)
model.load_state_dict(torch.load(trained_model_weights_path))
model.eval()

def extract_text_from_pdf(file_path):
    with pdfplumber.open(file_path) as pdf:
        text=''
        for page in pdf.pages:
            text+=page.extract_text()+'\n'
    return text

#用正则表达式实现多个目标的分隔
def split_sentences(text):
    sentences =re.split(r'[。！？]\s*',text)
    return [s for s in sentences if s]

#判断一堆句子中有无模式串
def filter_sentences_with_number(sentences):
    pattern=re.compile(r'\d+')
    return [s for s in sentences if pattern.search(s)]

def model_inference(sentences):
    length=len(sentences)
    positive=0
    negative=0
    for sentence in sentences:
        inputs=tokenizer(sentence,padding=True,truncation=True,return_tensors='pt')
        with torch.no_grad():
            outputs=model(**inputs)
        logits=outputs.logits
        prediction=torch.argmax(logits)
        if prediction==1:
            positive+=1
        else:
            negative+=1
    if positive>5*negative:
        label=1
    else:
        label=0
    return list([length,positive,negative,label])

#处理每一个文件的流程
def one_file_process(file_path,excel_path):
    text= extract_text_from_pdf(file_path)
    sentences = split_sentences(text)
    sentences = filter_sentences_with_number(sentences)
    result_list=model_inference(sentences)
    #向工作本中插入数据
    wb=openpyxl.load_workbook(excel_path)
    ws=wb[ws_title]
    max_row=ws.max_row
    ws.append(result_list)
    wb.save(excel_path)

def all_process(dir_path,excel_path):
    dir_path=os.path.abspath(dir_path)
    files=os.listdir(dir_path)
    for file in files:
        one_file_process(os.path.join(dir_path,file),excel_path)

#新建一个工作本
wb=openpyxl.Workbook()
ws=wb.active
ws.title=ws_title
excel_path=excel_path
wb.save(excel_path)
all_process(dir_path,excel_path)